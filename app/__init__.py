
# =========================
# IMPORTACIONES
# =========================
from flask import Flask, render_template, request, send_file, session, redirect, url_for, jsonify, flash
from flask_sqlalchemy import SQLAlchemy
import os
from io import BytesIO
from openpyxl.styles import Font, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl import Workbook
from reportlab.lib.pagesizes import letter
from reportlab.pdfgen import canvas
import webbrowser
from datetime import datetime, timedelta
import threading
from datetime import datetime, timedelta

# =========================
# CONSTANTES Y CONFIGURACIONES
# =========================
# Definición de constantes globales para masas base por 94 canastos
MASA_POR_94_CANASTOS = {
    'soja_kg': 50,
    'harina_kg': 39,
    'chimichurri_g': 800,
    'sal_g': 500
}
CANASTOS_BASE = 94

UNIDADES_POR_CANASTO = 18
MASA_BASE_POR_100_CANASTOS = {
    'soja_kg': 50,
    'chimichurri_g': 600,
    'sal_g': 500
}
MASA_BASE_CANASTOS = 100

MASA_ORIGINALES_POR_85_CANASTOS = {
    'soja_kg': 75,
    'harina_kg': 30,
    'chimichurri_g': 800,
    'sal_g': 500
}
MASA_ORIGINALES_CANASTOS = 85

# =========================
# INICIALIZACIÓN DE FLASK Y SQLALCHEMY
# =========================
app = Flask(__name__)
app.secret_key = 'alkimyk_clave_segura'
app.config['SQLALCHEMY_DATABASE_URI'] = 'postgresql://alkimyk_db_user:7vP5jvsKt9KEM8f9JZsd2dSdWGjiCphv@dpg-d0kfn13uibrs739gn9bg-a.oregon-postgres.render.com/alkimyk_db'
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False
db = SQLAlchemy(app)

# =========================
# FILTROS JINJA2
# =========================
# Filtro de plantilla para formatear fechas en los templates Jinja2
@app.template_filter('datetimeformat')
def datetimeformat(value, format='%A'):
    from datetime import datetime
    return datetime.strptime(value, '%Y-%m-%d').strftime(format)

# Filtro de plantilla para convertir strings de fecha a objetos datetime
@app.template_filter('to_datetime')
def to_datetime_filter(value, format='%Y-%m-%d'):
    from datetime import datetime
    return datetime.strptime(value, format)

# Filtro de plantilla para formatear números con formato argentino (1.234,56)
@app.template_filter('formato_argentino')
def formato_argentino(value):
    try:
        from babel.numbers import format_decimal
        value = float(value)
        if value.is_integer():
            value = int(value)
        return format_decimal(value, locale='es_AR')
    except:
        return value

# Filtro de plantilla para slugify en Jinja templates
@app.template_filter('slugify')
def slugify_filter(nombre):
    return nombre.strip().lower().replace(' ', '_').replace('á', 'a').replace('é', 'e').replace('í', 'i').replace('ó', 'o').replace('ú', 'u').replace('ñ', 'n')

# =========================
# MODELOS DE BASE DE DATOS
# =========================
# Modelo de usuario
class Usuario(db.Model):
    __tablename__ = 'usuarios'
    id = db.Column(db.Integer, primary_key=True)
    nombre = db.Column(db.String(100), nullable=False)
    email = db.Column(db.String(100), unique=True, nullable=False)
    password = db.Column(db.String(200), nullable=False)
    rol = db.Column(db.String(50), nullable=False)


# Modelo Produccion para guardar canastos por usuario
class Produccion(db.Model):
    __tablename__ = 'produccion'
    id = db.Column(db.Integer, primary_key=True)
    usuario_email = db.Column(db.String(100), nullable=False)
    fecha = db.Column(db.DateTime, default=datetime.utcnow)
    sabor = db.Column(db.String(50), nullable=False)
    canastos = db.Column(db.Integer, nullable=False)

# Modelo CostoFijo para guardar costos fijos por usuario
class CostoFijo(db.Model):
    __tablename__ = 'costos_fijos'
    id = db.Column(db.Integer, primary_key=True)
    usuario_email = db.Column(db.String(100), nullable=False)
    nombre = db.Column(db.String(100), nullable=False)
    monto = db.Column(db.Float, nullable=False)

# Modelo para guardar histórico de rentabilidad
class ResumenHistorico(db.Model):
    __tablename__ = 'resumen_historico'
    id = db.Column(db.Integer, primary_key=True)
    usuario_email = db.Column(db.String(100), nullable=False)
    fecha = db.Column(db.DateTime, default=datetime.utcnow)
    total_canastos = db.Column(db.Integer)
    total_cajas = db.Column(db.Integer)
    total_facturar = db.Column(db.Float)
    total_con_iva = db.Column(db.Float)
    ganancia_total = db.Column(db.Float)
    rentabilidad = db.Column(db.Float)

# Modelo para guardar precios de ingredientes por usuario
class PrecioIngrediente(db.Model):
    __tablename__ = 'precios_ingredientes'
    id = db.Column(db.Integer, primary_key=True)
    usuario_email = db.Column(db.String(100), nullable=False)
    ingrediente = db.Column(db.String(100), nullable=False)
    precio_unitario = db.Column(db.Float, nullable=False)

# Modelo para guardar precios de venta por sabor
class PrecioVentaSabor(db.Model):
    __tablename__ = 'precios_venta_sabor'
    id = db.Column(db.Integer, primary_key=True)
    usuario_email = db.Column(db.String(100), nullable=False)
    sabor = db.Column(db.String(100), nullable=False)
    precio = db.Column(db.Float, nullable=False)

class Proveedor(db.Model):
    __tablename__ = 'proveedores'
    id = db.Column(db.Integer, primary_key=True)
    nombre = db.Column(db.String(100), nullable=False)

# Modelo para guardar compras de ingredientes
class Compra(db.Model):
    __tablename__ = 'compras'
    id = db.Column(db.Integer, primary_key=True)
    ingrediente = db.Column(db.String(100), nullable=False)
    cantidad = db.Column(db.Float, nullable=False)
    proveedor = db.Column(db.String(100), nullable=False)
    forma_pago = db.Column(db.String(50), nullable=False)
    fecha_pago = db.Column(db.Date, nullable=False)

    # Modelo para guardar producción diaria
class ProduccionDiaria(db.Model):
    __tablename__ = 'produccion_diaria'
    id = db.Column(db.Integer, primary_key=True)
    fecha = db.Column(db.Date, nullable=False)
    sabor = db.Column(db.String(100), nullable=False)
    cantidad_canastos = db.Column(db.Integer, nullable=False)

# =========================
# FUNCIONES AUXILIARES
# =========================
# -----------------------------------------------------------
# Función para calcular el total de ingredientes (para planificación)
# -----------------------------------------------------------
def calcular_total_ingredientes(canastos):
    total_ingredientes = {}

    def add(dic, nombre, cantidad_g):
        dic[nombre] = dic.get(nombre, 0) + cantidad_g

    non_original_canastos = sum(c for s, c in canastos.items() if s != 'original')
    if non_original_canastos > 0:
        add(total_ingredientes, 'Soja', MASA_BASE_POR_100_CANASTOS['soja_kg'] * non_original_canastos / MASA_BASE_CANASTOS)
        add(total_ingredientes, 'Chimichurri', MASA_BASE_POR_100_CANASTOS['chimichurri_g'] * non_original_canastos / MASA_BASE_CANASTOS)
        add(total_ingredientes, 'Sal', MASA_BASE_POR_100_CANASTOS['sal_g'] * non_original_canastos / MASA_BASE_CANASTOS)

    for sabor, cantidad in canastos.items():
        if cantidad == 0:
            continue
        unidades = cantidad * (32 if sabor == 'original' else UNIDADES_POR_CANASTO)
        if sabor == 'original':
            receta_originales = {
                'Soja': 75,
                'Harina': 30,
                'Sal': 0.5,
                'Chimichurri': 0.8
            }
            for ingrediente, total_base in receta_originales.items():
                cantidad_ingrediente = (total_base / 85) * cantidad
                if ingrediente in ['Soja', 'Harina']:
                    cantidad_final = round(cantidad_ingrediente, 2)
                else:
                    cantidad_final = round(cantidad_ingrediente * 1000, 2) if cantidad_ingrediente < 1 else round(cantidad_ingrediente, 2)
                add(total_ingredientes, ingrediente, cantidad_final)
        else:
            if sabor == 'aceituna':
                add(total_ingredientes, 'Muzzarella', unidades * 15)
                add(total_ingredientes, 'Aceitunas', unidades * 20)
            elif sabor == 'caprese':
                tomate_total = unidades * 25
                add(total_ingredientes, 'Muzzarella', unidades * 15)
                add(total_ingredientes, 'Tomate', tomate_total)
                add(total_ingredientes, 'Albahaca', unidades * 2)
                add(total_ingredientes, 'Sal', (tomate_total / 1000) * 4)
            elif sabor == 'queso_azul':
                mezcla_total = unidades * 30
                porc_queso = 2.3 / (18 + 2.3)
                porc_muzza = 1 - porc_queso
                add(total_ingredientes, 'Muzzarella', mezcla_total * porc_muzza)
                add(total_ingredientes, 'Queso Azul', mezcla_total * porc_queso)
            elif sabor == 'cebolla':
                cebolla_cruda = (unidades * 40) / 0.8
                add(total_ingredientes, 'Cebolla', cebolla_cruda)
                add(total_ingredientes, 'Orégano', (cebolla_cruda / 1000) * 2)
                add(total_ingredientes, 'Sal', (cebolla_cruda / 1000) * 5)
            elif sabor == 'espinaca':
                total_relleno = unidades * 40 / 0.9
                espinaca = total_relleno * 0.5 / 0.9
                cebolla = total_relleno * 0.25 / 0.8
                morron = total_relleno * 0.25 / 0.8
                add(total_ingredientes, 'Espinaca', espinaca)
                add(total_ingredientes, 'Cebolla', cebolla)
                add(total_ingredientes, 'Morrón', morron)
                add(total_ingredientes, 'Nuez Moscada', total_relleno / 1000 * 1)
                add(total_ingredientes, 'Pimienta Negra', total_relleno / 1000 * 1)
                add(total_ingredientes, 'Sal', total_relleno / 1000 * 5)
            elif sabor == 'calabaza':
                total_relleno = unidades * 40 / 0.8
                add(total_ingredientes, 'Calabaza', total_relleno)
                add(total_ingredientes, 'Cúrcuma', total_relleno / 1000 * 5)
                add(total_ingredientes, 'Sal', total_relleno / 1000 * 5)
            elif sabor == 'brocoli':
                total_relleno = unidades * 40
                add(total_ingredientes, 'Brócoli', total_relleno * 0.6)
                add(total_ingredientes, 'Cebolla', total_relleno * 0.4 / 0.8)
                add(total_ingredientes, 'Chimichurri', total_relleno / 1000 * 5)
                add(total_ingredientes, 'Sal', total_relleno / 1000 * 5)

    for sabor, cantidad in canastos.items():
        if cantidad == 0:
            continue
        unidades_por_canasto = 32 if sabor == 'original' else UNIDADES_POR_CANASTO
        total_unidades = cantidad * unidades_por_canasto
        gramos_pan_rallado = total_unidades * 10
        add(total_ingredientes, 'Pan Rallado', gramos_pan_rallado)

    return total_ingredientes

def normalizar_importe(valor):
    try:
        if isinstance(valor, str):
            valor = valor.strip().replace('$', '')
            if valor == '':
                return 0
        # Convertir a float y devolver sin redondear ni formatear con ceros
        return float(valor)
    except Exception:
        return 0



# =========================
# RUTAS FLASK
# =========================



@app.route('/agregar_proveedor', methods=['POST'])
def agregar_proveedor():
    data = request.get_json()
    nombre = data.get('nombre')
    if not nombre:
        return jsonify({'success': False, 'message': 'Falta el nombre del proveedor'}), 400

    nuevo = Proveedor(nombre=nombre)
    db.session.add(nuevo)
    db.session.commit()

    return jsonify({'success': True, 'id': nuevo.id, 'nombre': nuevo.nombre, 'message': 'Proveedor guardado con éxito'})

# Ruta para registrar compras de ingredientes y devolver cantidad faltante
def registrar_compra():
    if 'usuario' not in session:
        return jsonify({'success': False, 'message': 'No autenticado'})

    ingrediente = request.form.get('ingrediente')
    cantidad = request.form.get('cantidad')
    proveedor = request.form.get('proveedor')
    forma_pago = request.form.get('forma_pago')
    fecha_pago = request.form.get('fecha_pago')

    if not ingrediente or not cantidad or not proveedor or not forma_pago or not fecha_pago:
        return jsonify({'success': False, 'message': 'Todos los campos son obligatorios'})

    try:
        cantidad = float(cantidad)
        if cantidad <= 0:
            raise ValueError
    except:
        return jsonify({'success': False, 'message': 'La cantidad debe ser válida'})

    nueva = Compra(
        ingrediente=ingrediente,
        cantidad=cantidad,
        proveedor=proveedor,
        forma_pago=forma_pago,
        fecha_pago=datetime.strptime(fecha_pago, '%Y-%m-%d')
    )
    db.session.add(nueva)
    db.session.commit()

    # Calcular cantidad restante actual
    from sqlalchemy import func
    total_comprado = db.session.query(func.sum(Compra.cantidad)).filter_by(ingrediente=ingrediente).scalar() or 0

    canastos = session.get('canastos', {})
    total_ingredientes = {}

    # Lógica mínima para sumar lo necesario para el ingrediente
    # Reutiliza código según lo usado en `planificacion`
    unidades_por_canasto = 32 if ingrediente.lower() == 'pan rallado' else UNIDADES_POR_CANASTO
    for sabor, cantidad_sabor in canastos.items():
        unidades = cantidad_sabor * unidades_por_canasto
        if ingrediente.lower() == 'pan rallado':
            total_ingredientes[ingrediente] = total_ingredientes.get(ingrediente, 0) + unidades * 10

    requerido = total_ingredientes.get(ingrediente, 0)
    if ingrediente not in ['Soja', 'Harina']:
        requerido = requerido / 1000 if requerido >= 1000 else requerido
    restante = max(0, round(requerido - total_comprado, 2))

    return jsonify({'success': True, 'message': 'Compra guardada', 'restante': restante})

# Nueva ruta para guardar compras completas desde el popup de planificación
@app.route('/guardar_compra', methods=['POST'])
def guardar_compra():
    if 'usuario' not in session:
        return jsonify({'success': False, 'message': 'No autenticado'})

    data = request.get_json()
    ingrediente = data.get('ingrediente')
    cantidad = data.get('cantidad')
    proveedor = data.get('proveedor')
    forma_pago = data.get('forma_pago')
    fecha_pago = data.get('fecha_pago')

    if not ingrediente or not cantidad or not proveedor or not forma_pago or not fecha_pago:
        return jsonify({'success': False, 'message': 'Todos los campos son obligatorios'})

    try:
        cantidad = float(cantidad)
        if cantidad <= 0:
            raise ValueError
    except:
        return jsonify({'success': False, 'message': 'La cantidad debe ser válida'})

    nueva = Compra(
        ingrediente=ingrediente,
        cantidad=cantidad,
        proveedor=proveedor,
        forma_pago=forma_pago,
        fecha_pago=datetime.strptime(fecha_pago, '%Y-%m-%d')
    )
    db.session.add(nueva)
    db.session.commit()

    return jsonify({'success': True, 'message': 'Compra registrada correctamente'})

@app.route('/')
def home():
    if 'usuario' in session and session.get('rol') == 'admin':
        return render_template('index.html')
    else:
        return redirect(url_for('login_admin'))

@app.route('/canastos', methods=['GET', 'POST'])
def canastos():
    ingredientes = {}
    detalles_por_sabor = {}
    mostrar = False
    canastos = {}
    errores = []

    # Si el método es GET y hay datos en sesión, mostrar los cuadros
    if request.method == 'GET' and 'canastos' in session:
        canastos = session['canastos']
        mostrar = True

    # Si el método es POST, validar y calcular
    if request.method == 'POST':
        # --- Manejo del botón "Limpiar"
        if 'limpiar' in request.form:
            session.pop('canastos', None)
            session.pop('cupo_diario', None)
            canastos = {
                'aceituna': 0,
                'caprese': 0,
                'queso_azul': 0,
                'cebolla': 0,
                'espinaca': 0,
                'calabaza': 0,
                'brocoli': 0
            }
            session['limpiar'] = True
            return render_template('canastos.html',
                                   mostrar=False,
                                   canastos=canastos,
                                   detalles_por_sabor={},
                                   ingredientes={},
                                   total_cajas=0,
                                   cupo_diario_default=110,
                                   dias_produccion=0,
                                   errores=[],
                                   limpiar=True)
        # --- Fin manejo limpiar
        canastos = {}
        datos_formulario = {k: v.strip() for k, v in request.form.items()}
        for sabor, valor in datos_formulario.items():
            if sabor == "cupo_diario":
                continue  # No tratar este campo como un sabor
            if not valor:
                errores.append(f"El campo {sabor.capitalize()} no puede estar vacío.")
            elif not valor.isdigit():
                errores.append(f"Debe ingresar un número válido para {sabor.capitalize()}.")
            else:
                canastos[sabor] = int(valor)
        if errores:
            mostrar = False
            return render_template('canastos.html', errores=errores, mostrar=mostrar, canastos=datos_formulario)
        mostrar = True
        session['canastos'] = canastos
        # Guardar canastos en la tabla Produccion por usuario
        if 'usuario' in session:
            usuario_email = session['usuario']
            Produccion.query.filter_by(usuario_email=usuario_email).delete()
            for sabor, cantidad in canastos.items():
                if cantidad > 0:
                    nueva_prod = Produccion(usuario_email=usuario_email, sabor=sabor, canastos=cantidad)
                    db.session.add(nueva_prod)
            db.session.commit()
        total_canastos = sum(canastos.values())

        # Definir unidades por canasto para original
        UNIDADES_POR_CANASTO_ORIGINAL = 32

        total_ingredientes = {}
        # Para sabores que NO sean 'original'
        non_original_canastos = sum(c for s, c in canastos.items() if s != 'original')
        def add(dic, nombre, cantidad_g):
            dic[nombre] = dic.get(nombre, 0) + cantidad_g
        if non_original_canastos > 0:
            soja_no_original = MASA_BASE_POR_100_CANASTOS['soja_kg'] * non_original_canastos / MASA_BASE_CANASTOS
            chimi_no_original = MASA_BASE_POR_100_CANASTOS['chimichurri_g'] * non_original_canastos / MASA_BASE_CANASTOS
            sal_no_original = MASA_BASE_POR_100_CANASTOS['sal_g'] * non_original_canastos / MASA_BASE_CANASTOS

            add(total_ingredientes, 'Soja', soja_no_original)
            add(total_ingredientes, 'Chimichurri', chimi_no_original)
            add(total_ingredientes, 'Sal', sal_no_original)
            harina_no_original = MASA_POR_94_CANASTOS['harina_kg'] * non_original_canastos / CANASTOS_BASE
            add(total_ingredientes, 'Harina', harina_no_original)

        # Agregar receta de Original si existe
        if 'original' in canastos and canastos['original'] > 0:
            cantidad = canastos['original']
            receta_originales = {
                'Soja': 75,
                'Harina': 30,
                'Sal': 0.5,
                'Chimichurri': 0.8
            }
            detalles_por_sabor['original'] = {}
            for ingrediente, total_base in receta_originales.items():
                cantidad_ingrediente = (total_base / 85) * cantidad
                if ingrediente in ['Soja', 'Harina']:
                    cantidad_final = round(cantidad_ingrediente, 2)
                    detalles_por_sabor['original'][ingrediente] = cantidad_final
                    # Sumar SIEMPRE en kg para Soja y Harina
                    total_ingredientes[ingrediente] = total_ingredientes.get(ingrediente, 0) + cantidad_ingrediente
                else:
                    cantidad_final = round(cantidad_ingrediente * 1000, 2) if cantidad_ingrediente < 1 else round(cantidad_ingrediente, 2)
                    detalles_por_sabor['original'][ingrediente] = cantidad_final
                    # Sumar en g para los demás
                    total_ingredientes[ingrediente] = total_ingredientes.get(ingrediente, 0) + cantidad_final

        for sabor, cantidad in canastos.items():
            if cantidad == 0:
                continue
            # No duplicar la lógica para original, ya está arriba
            if sabor == 'original':
                continue
            unidades = cantidad * UNIDADES_POR_CANASTO
            detalles_por_sabor[sabor] = {}
            temp = detalles_por_sabor[sabor]
            # Masa base NO se agrega al desglose por sabor (solo ingredientes específicos del relleno)
            if sabor == 'aceituna':
                add(temp, 'Muzzarella', unidades * 15)
                add(temp, 'Aceitunas', unidades * 20)
            elif sabor == 'caprese':
                tomate_total = unidades * 25
                add(temp, 'Muzzarella', unidades * 15)
                add(temp, 'Tomate', tomate_total)
                add(temp, 'Albahaca', unidades * 2)
                add(temp, 'Sal', (tomate_total / 1000) * 4)
            elif sabor == 'queso_azul':
                mezcla_total = unidades * 30
                porc_queso = 2.3 / (18 + 2.3)
                porc_muzza = 1 - porc_queso
                add(temp, 'Muzzarella', mezcla_total * porc_muzza)
                add(temp, 'Queso Azul', mezcla_total * porc_queso)
            elif sabor == 'cebolla':
                cebolla_cruda = (unidades * 40) / 0.8
                add(temp, 'Cebolla', cebolla_cruda)
                add(temp, 'Orégano', (cebolla_cruda / 1000) * 2)
                add(temp, 'Sal', (cebolla_cruda / 1000) * 5)
            elif sabor == 'espinaca':
                total_relleno = unidades * 40 / 0.9
                espinaca = total_relleno * 0.5 / 0.9
                cebolla = total_relleno * 0.25 / 0.8
                morron = total_relleno * 0.25 / 0.8
                add(temp, 'Espinaca', espinaca)
                add(temp, 'Cebolla', cebolla)
                add(temp, 'Morrón', morron)
                add(temp, 'Nuez Moscada', total_relleno / 1000 * 1)
                add(temp, 'Pimienta Negra', total_relleno / 1000 * 1)
                add(temp, 'Sal', total_relleno / 1000 * 5)
            elif sabor == 'calabaza':
                total_relleno = unidades * 40 / 0.8
                add(temp, 'Calabaza', total_relleno)
                add(temp, 'Cúrcuma', total_relleno / 1000 * 5)
                add(temp, 'Sal', total_relleno / 1000 * 5)
            elif sabor == 'brocoli':
                total_relleno = unidades * 40
                add(temp, 'Brócoli', total_relleno * 0.6)
                add(temp, 'Cebolla', total_relleno * 0.4 / 0.8)
                add(temp, 'Chimichurri', total_relleno / 1000 * 5)
                add(temp, 'Sal', total_relleno / 1000 * 5)
            for k, v in temp.items():
                total_ingredientes[k] = total_ingredientes.get(k, 0) + v

        ingredientes = total_ingredientes
        # Nuevo cálculo de total_cajas para 'original'
        total_cajas = 0
        for sabor, cantidad in canastos.items():
            unidades_por_canasto = 32 if sabor == 'original' else UNIDADES_POR_CANASTO
            total_unidades = cantidad * unidades_por_canasto
            if sabor == 'original':
                total_cajas += round(total_unidades / 108)
            else:
                total_cajas += round(total_unidades / (15 * 4))

        # Obtener el valor por defecto del límite diario de producción
        cupo_diario = request.form.get('cupo_diario')
        if cupo_diario and cupo_diario.isdigit():
            session['cupo_diario'] = int(cupo_diario)
            cupo_diario_default = int(cupo_diario)
        else:
            cupo_diario_default = session.get('cupo_diario', 110)

        # Calcular la cantidad de días de producción según el cupo diario
        dias_produccion = (total_canastos + cupo_diario_default - 1) // cupo_diario_default
    else:
        # Si es GET o no se envió POST, preparar valores por defecto
        total_cajas = 0
        if canastos:
            for sabor, cantidad in canastos.items():
                unidades_por_canasto = 32 if sabor == 'original' else UNIDADES_POR_CANASTO
                total_unidades = cantidad * unidades_por_canasto
                if sabor == 'original':
                    total_cajas += round(total_unidades / 108)
                else:
                    total_cajas += round(total_unidades / (15 * 4))
        cupo_diario = request.form.get('cupo_diario')
        if cupo_diario and cupo_diario.isdigit():
            session['cupo_diario'] = int(cupo_diario)
            cupo_diario_default = int(cupo_diario)
        else:
            cupo_diario_default = session.get('cupo_diario', 110)
        total_canastos = sum(canastos.values()) if canastos else 0
        dias_produccion = (total_canastos + cupo_diario_default - 1) // cupo_diario_default if total_canastos > 0 else 0

    # Agregar pan rallado (10 g por unidad)
    for sabor, cantidad in canastos.items():
        if cantidad == 0:
            continue
        unidades_por_canasto = 32 if sabor == 'original' else UNIDADES_POR_CANASTO
        total_unidades = cantidad * unidades_por_canasto
        gramos_pan_rallado = total_unidades * 10  # 10 gramos por unidad
        detalles_por_sabor.setdefault(sabor, {})
        detalles_por_sabor[sabor]['Pan Rallado'] = gramos_pan_rallado
        ingredientes['Pan Rallado'] = ingredientes.get('Pan Rallado', 0) + gramos_pan_rallado
     

    # Obtener los días seleccionados de la configuración para mostrar en la plantilla
    dias_seleccionados = session.get('dias_habilitados', ['lunes', 'martes', 'miércoles', 'jueves', 'viernes'])
    mostrar_boton_costos = 'usuario' in session and session.get('rol') == 'admin' and bool(ingredientes)
    # --- Agregar prints para debug de datos totalizados ---
    print("📦 Canastos finales:", canastos)
    print("🧮 Ingredientes totales:", ingredientes)
    print("📋 Detalles por sabor:", detalles_por_sabor)
    print("📦 Total de cajas:", total_cajas)
    print("📅 Días de producción:", dias_produccion)
    # ------------------------------------------------------
    # Guardar datos de planificación en la sesión (para /planificacion y otros)
    session['planificacion'] = True
    session['total_ingredientes_fmt'] = {
    k: {
    'cantidad': round(v, 2) if k in ['Soja', 'Harina'] else (round(v / 1000, 2) if v >= 1000 else round(v, 2)),
    'unidad': 'kg' if k in ['Soja', 'Harina'] or v >= 1000 else 'g'
    }
    for k, v in ingredientes.items()
    }
    session['total_canastos'] = sum(canastos.values())
    session['total_cajas'] = total_cajas
    # --- AGREGADO: guardar canastos en sesión para planificación ---
    session['canastos'] = canastos
    return render_template('canastos.html',
                           ingredientes=ingredientes,
                           mostrar=mostrar,
                           canastos=canastos,
                           detalles_por_sabor=detalles_por_sabor,
                           total_cajas=total_cajas,
                           cupo_diario_default=cupo_diario_default,
                           dias_produccion=dias_produccion,
                           dias_seleccionados=dias_seleccionados,
                           mostrar_boton_costos=mostrar_boton_costos)




from flask import jsonify, request, session
from datetime import datetime, timedelta

@app.route('/generar_calendario', methods=['POST'])
def generar_calendario():
    fecha_inicio = request.form.get('fecha_inicio')
    if not fecha_inicio:
        return "Fecha de inicio no proporcionada", 400

    # Recuperar cupo_diario del formulario o de la sesión
    cupo_diario = request.form.get('cupo_diario', None)
    if cupo_diario is not None:
        try:
            cupo_diario = int(cupo_diario)
        except ValueError:
            cupo_diario = 110
        session['cupo_diario'] = cupo_diario
    else:
        cupo_diario = session.get('cupo_diario', 110)

    try:
        fecha = datetime.strptime(fecha_inicio, '%Y-%m-%d')
    except ValueError:
        return "Formato de fecha inválido", 400

    canastos = session.get('canastos', {}).copy()
    total = sum(canastos.values())
    dias = []
    producidos = 0

    dias_habilitados = session.get('dias_habilitados', [])
    dias_habilitados = [d.lower() for d in dias_habilitados]

    dias_traducidos = {
        'monday': 'lunes',
        'tuesday': 'martes',
        'wednesday': 'miércoles',
        'thursday': 'jueves',
        'friday': 'viernes',
        'saturday': 'sábado',
        'sunday': 'domingo'
    }
    while producidos < total:
        dia_semana_en = fecha.strftime('%A').lower()
        dia_semana_es = dias_traducidos.get(dia_semana_en, '').lower()
        if dia_semana_es not in dias_habilitados:
            fecha += timedelta(days=1)
            continue

        restante = total - producidos
        hoy = min(cupo_diario, restante)
        sabores_distribuidos = {}
        restantes = hoy
        # Distribuir los canastos entre sabores
        ranking_sabores = ['caprese', 'queso_azul', 'espinaca', 'aceituna', 'calabaza', 'brocoli', 'cebolla']
        for sabor in canastos.keys():
            # preserve existing order if not in ranking
            if sabor not in ranking_sabores:
                ranking_sabores.append(sabor)
        for sabor in ranking_sabores:
            cantidad = canastos.get(sabor, 0)
            if cantidad <= 0:
                continue
            asignar = min(cantidad, restantes)
            if asignar > 0:
                sabores_distribuidos[sabor] = asignar
                canastos[sabor] -= asignar
                restantes -= asignar
            if restantes == 0:
                break

        # Reiniciar hoy a 0 antes del bloque de asignación forzada
        hoy = 0
        # Asegura que haya al menos un sabor asignado si quedan canastos de algún sabor
        if not sabores_distribuidos and any(c > 0 for c in canastos.values()):
            for sabor in ranking_sabores:
                if canastos.get(sabor, 0) > 0:
                    sabores_distribuidos[sabor] = 1
                    canastos[sabor] -= 1
                    hoy = 1
                    break

        # Solo agregar el día si hay algún sabor distribuido
        if sabores_distribuidos:
            dias.append({
                "fecha": fecha.strftime('%Y-%m-%d'),
                "canastos": sum(sabores_distribuidos.values()),
                "sabores": sabores_distribuidos
            })
            producidos += sum(sabores_distribuidos.values())
        fecha += timedelta(days=1)

    return jsonify(dias)

# Ruta para generación y visualización del calendario de producción
@app.route('/calendario', methods=['GET', 'POST'])
def calendario():
    from datetime import datetime, timedelta

    if request.method == 'POST':
        canastos = session.get('canastos', {})
        if not canastos:
            return redirect(url_for('canastos'))

        cupo_diario = request.form.get('cupo_diario')
        try:
            cupo_diario = int(cupo_diario)
        except (TypeError, ValueError):
            cupo_diario = session.get('cupo_diario', 110)
        session['cupo_diario'] = cupo_diario

        fecha_inicio_str = request.form.get('fecha_inicio')
        if not fecha_inicio_str:
            return redirect(url_for('canastos'))

        fecha_actual = datetime.strptime(fecha_inicio_str, '%Y-%m-%d').date()
        calendario = []
        canastos_restantes = sum(canastos.values())
        canastos_sabores = canastos.copy()

        dias_habilitados = session.get('dias_habilitados', ['lunes', 'martes', 'miércoles', 'jueves', 'viernes'])
        dias_habilitados = [d.lower() for d in dias_habilitados]
        dias_traducidos = {
            'monday': 'lunes',
            'tuesday': 'martes',
            'wednesday': 'miércoles',
            'thursday': 'jueves',
            'friday': 'viernes',
            'saturday': 'sábado',
            'sunday': 'domingo'
        }

        for _ in range(30):
            dia_semana_en = fecha_actual.strftime('%A').lower()
            dia_semana_es = dias_traducidos.get(dia_semana_en, '')

            dia_info = {
                'fecha': fecha_actual.strftime('%Y-%m-%d'),
                'dia_semana': dia_semana_es,
                'canastos': 0,
                'sabores': {}
            }

            if dia_semana_es != 'lunes' and dia_semana_es in dias_habilitados and canastos_restantes > 0:
                sabores_distribuidos = {}
                restantes = min(canastos_restantes, cupo_diario)
                ranking_sabores = ['caprese', 'queso_azul', 'espinaca', 'aceituna', 'calabaza', 'brocoli', 'cebolla', 'original']

                for sabor in ranking_sabores:
                    cantidad = canastos_sabores.get(sabor, 0)
                    if cantidad <= 0:
                        continue
                    asignar = min(cantidad, restantes)
                    if asignar > 0:
                        sabores_distribuidos[sabor] = asignar
                        canastos_sabores[sabor] -= asignar
                        restantes -= asignar
                    if restantes == 0:
                        break

                if not sabores_distribuidos and any(c > 0 for c in canastos_sabores.values()):
                    for sabor in ranking_sabores:
                        if canastos_sabores.get(sabor, 0) > 0:
                            sabores_distribuidos[sabor] = 1
                            canastos_sabores[sabor] -= 1
                            break

                if sabores_distribuidos:
                    dia_info['canastos'] = sum(sabores_distribuidos.values())
                    dia_info['sabores'] = sabores_distribuidos
                    canastos_restantes -= dia_info['canastos']

            calendario.append(dia_info)
            fecha_actual += timedelta(days=1)

        cards = []
        for dia in calendario:
            fecha = dia['fecha']
            for sabor, cantidad in dia['sabores'].items():
                cards.append({
                    'sabor': sabor.capitalize(),
                    'cantidad': cantidad,
                    'fecha': fecha
                })

        return render_template('calendario.html', calendario=calendario, fecha_inicio=fecha_inicio_str, timedelta=timedelta, cards=cards)

    # Si es GET
    hoy = datetime.today().strftime('%Y-%m-%d')
    sabores_totales = [
        'Caprese', 'Queso Azul', 'Aceituna', 'Espinaca', 'Brócoli',
        'Calabaza', 'Cebolla', 'Original'
    ]

    canastos = session.get('canastos', {})
    cards = []

    for sabor in sabores_totales:
        cantidad = canastos.get(sabor.lower(), 0) or canastos.get(sabor, 0)
        if cantidad > 0:
            cards.append({
                'sabor': sabor,
                'cantidad': cantidad,
                'fecha': hoy
            })

    return render_template('calendario.html', calendario=[], fecha_inicio=hoy, timedelta=timedelta, cards=cards)
@app.route('/stock', methods=['GET', 'POST'])
def stock():
    cajas = {}
    canastos = {}
    if request.method == 'POST':
        errores = []
        cajas = {}
        for sabor in ['aceituna', 'caprese', 'queso_azul', 'cebolla', 'espinaca', 'calabaza', 'brocoli', 'original']:
            valor = request.form.get(sabor, '').strip()
            if not valor:
                errores.append(f"El campo {sabor.capitalize()} no puede estar vacío.")
            elif not valor.isdigit():
                errores.append(f"Debe ingresar un número válido para {sabor.capitalize()}.")
            else:
                cajas[sabor] = int(valor)
        if errores:
            return render_template('stock.html', cajas=request.form, errores=errores)
        for sabor, cant_cajas in cajas.items():
            cant_cajas = int(cant_cajas) if cant_cajas else 0
            if sabor == 'original':
                total_unidades = cant_cajas * 27 * 4  # 27 packs por caja, 4 unidades por pack
                canastos[sabor] = round(total_unidades / 32)
            else:
                total_unidades = cant_cajas * 15 * 4  # 15 packs por caja para otros sabores
                canastos[sabor] = round(total_unidades / 18)
        session['canastos'] = canastos
        return redirect(url_for('canastos'))
    return render_template('stock.html', cajas=cajas, canastos=canastos)

@app.route('/exportar', methods=['POST'])
def exportar_excel():
    canastos = session.get('canastos', {})
    if not canastos:
        return "No hay datos para exportar", 400

    detalles_por_sabor = {}
    ingredientes_totales = {}
    total_canastos = sum(canastos.values())
    # Cálculo especial de cajas considerando "original" con 32 unidades por canasto y 108 unidades por caja
    total_cajas = 0
    for sabor, cantidad in canastos.items():
        unidades_por_canasto = 32 if sabor == 'original' else UNIDADES_POR_CANASTO
        total_unidades = cantidad * unidades_por_canasto
        if sabor == 'original':
            total_cajas += round(total_unidades / 108)
        else:
            total_cajas += round(total_unidades / (15 * 4))
    # Masa base para no-original
    non_original_canastos = sum(c for s, c in canastos.items() if s != 'original')
    if non_original_canastos > 0:
        ingredientes_totales['Soja'] = MASA_BASE_POR_100_CANASTOS['soja_kg'] * non_original_canastos / MASA_BASE_CANASTOS
        ingredientes_totales['Chimichurri'] = MASA_BASE_POR_100_CANASTOS['chimichurri_g'] * non_original_canastos / MASA_BASE_CANASTOS
        ingredientes_totales['Sal'] = MASA_BASE_POR_100_CANASTOS['sal_g'] * non_original_canastos / MASA_BASE_CANASTOS

    for sabor, cantidad in canastos.items():
        if cantidad == 0:
            continue
        unidades = cantidad * (32 if sabor == 'original' else UNIDADES_POR_CANASTO)
        temp = {}
        if sabor == 'original':
            receta_originales = {
                'Soja': 75,
                'Harina': 30,
                'Sal': 0.5,
                'Chimichurri': 0.8
            }
            for ingrediente, total_base in receta_originales.items():
                cantidad_ingrediente = (total_base / 85) * cantidad
                if ingrediente in ['Soja', 'Harina']:
                    cantidad_final = round(cantidad_ingrediente, 2)
                else:
                    cantidad_final = round(cantidad_ingrediente * 1000, 2) if cantidad_ingrediente < 1 else round(cantidad_ingrediente, 2)
                temp[ingrediente] = cantidad_final
                ingredientes_totales[ingrediente] = ingredientes_totales.get(ingrediente, 0) + cantidad_final
        else:
            # Masa base por sabor
            temp['Soja'] = MASA_BASE_POR_100_CANASTOS['soja_kg'] * cantidad / MASA_BASE_CANASTOS
            temp['Chimichurri'] = MASA_BASE_POR_100_CANASTOS['chimichurri_g'] * cantidad / MASA_BASE_CANASTOS
            temp['Sal'] = MASA_BASE_POR_100_CANASTOS['sal_g'] * cantidad / MASA_BASE_CANASTOS
            if sabor == 'aceituna':
                temp['Muzzarella'] = unidades * 15
                temp['Aceitunas'] = unidades * 20
            elif sabor == 'caprese':
                tomate_total = unidades * 25
                temp['Muzzarella'] = unidades * 15
                temp['Tomate'] = tomate_total
                temp['Albahaca'] = unidades * 2
                temp['Sal'] = temp.get('Sal', 0) + (tomate_total / 1000) * 4
            elif sabor == 'queso_azul':
                mezcla_total = unidades * 30
                porc_queso = 2.3 / (18 + 2.3)
                porc_muzza = 1 - porc_queso
                temp['Muzzarella'] = mezcla_total * porc_muzza
                temp['Queso Azul'] = mezcla_total * porc_queso
            elif sabor == 'cebolla':
                cebolla_cruda = (unidades * 40) / 0.8
                temp['Cebolla'] = cebolla_cruda
                temp['Orégano'] = (cebolla_cruda / 1000) * 2
                temp['Sal'] = temp.get('Sal', 0) + (cebolla_cruda / 1000) * 5
            elif sabor == 'espinaca':
                total_relleno = unidades * 40 / 0.9
                espinaca = total_relleno * 0.5 / 0.9
                cebolla = total_relleno * 0.25 / 0.8
                morron = total_relleno * 0.25 / 0.8
                temp['Espinaca'] = espinaca
                temp['Cebolla'] = cebolla
                temp['Morrón'] = morron
                temp['Nuez Moscada'] = total_relleno / 1000 * 1
                temp['Pimienta Negra'] = total_relleno / 1000 * 1
                temp['Sal'] = temp.get('Sal', 0) + total_relleno / 1000 * 5
            elif sabor == 'calabaza':
                total_relleno = unidades * 40 / 0.8
                temp['Calabaza'] = total_relleno
                temp['Cúrcuma'] = total_relleno / 1000 * 5
                temp['Sal'] = temp.get('Sal', 0) + total_relleno / 1000 * 5
            elif sabor == 'brocoli':
                total_relleno = unidades * 40
                temp['Brócoli'] = total_relleno * 0.6
                temp['Cebolla'] = total_relleno * 0.4 / 0.8
                temp['Chimichurri'] = total_relleno / 1000 * 5
                temp['Sal'] = temp.get('Sal', 0) + total_relleno / 1000 * 5
            for k, v in temp.items():
                ingredientes_totales[k] = ingredientes_totales.get(k, 0) + v
        detalles_por_sabor[sabor] = temp

    # Agregar pan rallado (10 g por unidad)
    for sabor, cantidad in canastos.items():
        if cantidad == 0:
            continue
        unidades_por_canasto = 32 if sabor == 'original' else UNIDADES_POR_CANASTO
        total_unidades = cantidad * unidades_por_canasto
        gramos_pan_rallado = total_unidades * 10  # 10 gramos por unidad
        detalles_por_sabor.setdefault(sabor, {})
        detalles_por_sabor[sabor]['Pan Rallado'] = gramos_pan_rallado
        ingredientes_totales['Pan Rallado'] = ingredientes_totales.get('Pan Rallado', 0) + gramos_pan_rallado

    output = BytesIO()
    wb = Workbook()

    # Ingredientes por sabor
    ws1 = wb.active
    ws1.title = "Ingredientes por sabor"
    font = Font(size=19)
    border = Border(left=Side(style='thin'), right=Side(style='thin'),
                    top=Side(style='thin'), bottom=Side(style='thin'))
    row = 1
    for sabor, ingredientes in detalles_por_sabor.items():
        if row > 1:
            row += 2
        cantidad_canastos = canastos[sabor]
        # Usar 32 unidades para 'original', si corresponde, y 108 unidades por caja
        if sabor == 'original':
            cantidad_cajas = round((cantidad_canastos * 32) / 108)
        else:
            cantidad_cajas = round((cantidad_canastos * UNIDADES_POR_CANASTO) / (15 * 4))
        ws1.cell(row=row, column=1, value=f"{sabor.replace('_', ' ').capitalize()} ({cantidad_canastos} canastos, {cantidad_cajas} cajas)").font = Font(size=19, bold=True)
        row += 1
        ws1.cell(row=row, column=1, value="Ingrediente").font = font
        ws1.cell(row=row, column=2, value="Cantidad").font = font
        ws1.cell(row=row, column=3, value="Unidad").font = font
        row += 1
        for ingr, cant in ingredientes.items():
            ws1.cell(row=row, column=1, value=ingr).font = font
            if cant >= 1000:
                ws1.cell(row=row, column=2, value=round(cant/1000, 2)).font = font
                ws1.cell(row=row, column=3, value="kg").font = font
            else:
                ws1.cell(row=row, column=2, value=round(cant, 2)).font = font
                ws1.cell(row=row, column=3, value="g").font = font
            row += 1
    for r in ws1.iter_rows(min_row=1, max_row=row, min_col=1, max_col=3):
        for cell in r:
            cell.border = border
    for col in range(1, 4):
        ws1.column_dimensions[get_column_letter(col)].auto_size = True

    # Total ingredientes
    ws2 = wb.create_sheet("Total Ingredientes")
    ws2.cell(row=1, column=1, value="Ingrediente").font = font
    ws2.cell(row=1, column=2, value="Cantidad").font = font
    ws2.cell(row=1, column=3, value="Unidad").font = font
    for i, (ingr, cant) in enumerate(ingredientes_totales.items(), start=2):
        ws2.cell(row=i, column=1, value=ingr).font = font
        if ingr in ['Soja', 'Harina']:
            ws2.cell(row=i, column=2, value=round(cant, 2)).font = font
            ws2.cell(row=i, column=3, value="kg").font = font
        else:
            if cant >= 1000:
                ws2.cell(row=i, column=2, value=round(cant/1000, 2)).font = font
                ws2.cell(row=i, column=3, value="kg").font = font
            else:
                ws2.cell(row=i, column=2, value=round(cant, 2)).font = font
                ws2.cell(row=i, column=3, value="g").font = font
    i += 2
    ws2.cell(row=i, column=1, value=f"Total de ingredientes para comprar y elaborar {total_canastos} canastos").font = font
    i += 1
    ws2.cell(row=i, column=1, value=f"Total de cajas: {total_cajas}").font = font
    for r in ws2.iter_rows(min_row=1, max_row=i, min_col=1, max_col=3):
        for cell in r:
            cell.border = border
    for col in range(1, 4):
        ws2.column_dimensions[get_column_letter(col)].auto_size = True

    wb.save(output)
    output.seek(0)
    return send_file(output, download_name='produccion_alkimyk.xlsx', as_attachment=True)

@app.route('/exportar_pdf', methods=['POST'])
def exportar_pdf():
    canastos = session.get('canastos', {})
    if not canastos:
        return "No hay datos para exportar", 400

    tipo = request.form.get('tipo')
    detalles_por_sabor = {}
    total_ingredientes = {}
    total_canastos = sum(canastos.values())
    # Cálculo especial de cajas considerando "original" con 32 unidades por canasto y 108 unidades por caja
    total_cajas = 0
    for sabor, cantidad in canastos.items():
        unidades = cantidad * (32 if sabor == 'original' else UNIDADES_POR_CANASTO)
        if sabor == 'original':
            total_cajas += round(unidades / 108)
        else:
            total_cajas += round(unidades / (15 * 4))

    for sabor, cantidad in canastos.items():
        if cantidad == 0:
            continue
        unidades = cantidad * (32 if sabor == 'original' else UNIDADES_POR_CANASTO)
        temp = {}
        if sabor == 'original':
            receta_originales = {
                'Soja': 75,
                'Harina': 30,
                'Sal': 0.5,
                'Chimichurri': 0.8
            }
            for ingrediente, total_base in receta_originales.items():
                cantidad_ingrediente = (total_base / 85) * cantidad
                if ingrediente in ['Soja', 'Harina']:
                    cantidad_final = round(cantidad_ingrediente, 2)
                else:
                    cantidad_final = round(cantidad_ingrediente * 1000, 2) if cantidad_ingrediente < 1 else round(cantidad_ingrediente, 2)
                temp[ingrediente] = cantidad_final
                total_ingredientes[ingrediente] = total_ingredientes.get(ingrediente, 0) + cantidad_final
        else:
            temp['Soja'] = MASA_BASE_POR_100_CANASTOS['soja_kg'] * cantidad / MASA_BASE_CANASTOS
            temp['Chimichurri'] = MASA_BASE_POR_100_CANASTOS['chimichurri_g'] * cantidad / MASA_BASE_CANASTOS
            temp['Sal'] = MASA_BASE_POR_100_CANASTOS['sal_g'] * cantidad / MASA_BASE_CANASTOS
            if sabor == 'aceituna':
                temp['Muzzarella'] = unidades * 15
                temp['Aceitunas'] = unidades * 20
            elif sabor == 'caprese':
                tomate_total = unidades * 25
                temp['Muzzarella'] = unidades * 15
                temp['Tomate'] = tomate_total
                temp['Albahaca'] = unidades * 2
                temp['Sal'] = temp.get('Sal', 0) + (tomate_total / 1000) * 4
            elif sabor == 'queso_azul':
                mezcla_total = unidades * 30
                porc_queso = 2.3 / (18 + 2.3)
                porc_muzza = 1 - porc_queso
                temp['Muzzarella'] = mezcla_total * porc_muzza
                temp['Queso Azul'] = mezcla_total * porc_queso
            elif sabor == 'cebolla':
                cebolla_cruda = (unidades * 40) / 0.8
                temp['Cebolla'] = cebolla_cruda
                temp['Orégano'] = (cebolla_cruda / 1000) * 2
                temp['Sal'] = temp.get('Sal', 0) + (cebolla_cruda / 1000) * 5
            elif sabor == 'espinaca':
                total_relleno = unidades * 40 / 0.9
                espinaca = total_relleno * 0.5 / 0.9
                cebolla = total_relleno * 0.25 / 0.8
                morron = total_relleno * 0.25 / 0.8
                temp['Espinaca'] = espinaca
                temp['Cebolla'] = cebolla
                temp['Morrón'] = morron
                temp['Nuez Moscada'] = total_relleno / 1000 * 1
                temp['Pimienta Negra'] = total_relleno / 1000 * 1
                temp['Sal'] = temp.get('Sal', 0) + total_relleno / 1000 * 5
            elif sabor == 'calabaza':
                total_relleno = unidades * 40 / 0.8
                temp['Calabaza'] = total_relleno
                temp['Cúrcuma'] = total_relleno / 1000 * 5
                temp['Sal'] = temp.get('Sal', 0) + total_relleno / 1000 * 5
            elif sabor == 'brocoli':
                total_relleno = unidades * 40
                temp['Brócoli'] = total_relleno * 0.6
                temp['Cebolla'] = total_relleno * 0.4 / 0.8
                temp['Chimichurri'] = total_relleno / 1000 * 5
                temp['Sal'] = temp.get('Sal', 0) + total_relleno / 1000 * 5
            for k, v in temp.items():
                total_ingredientes[k] = total_ingredientes.get(k, 0) + v
        detalles_por_sabor[sabor] = temp

    # Agregar pan rallado (10 g por unidad)
    for sabor, cantidad in canastos.items():
        if cantidad == 0:
            continue
        unidades_por_canasto = 32 if sabor == 'original' else UNIDADES_POR_CANASTO
        total_unidades = cantidad * unidades_por_canasto
        gramos_pan_rallado = total_unidades * 10  # 10 gramos por unidad
        detalles_por_sabor.setdefault(sabor, {})
        detalles_por_sabor[sabor]['Pan Rallado'] = gramos_pan_rallado
        total_ingredientes['Pan Rallado'] = total_ingredientes.get('Pan Rallado', 0) + gramos_pan_rallado

    output = BytesIO()
    pdf_output = output
    c = canvas.Canvas(pdf_output, pagesize=letter)
    width, height = letter
    try:
        c.drawImage("static/logo.png", width - 150, height - 80, width=100, preserveAspectRatio=True, mask='auto')
    except:
        pass  # Si el logo no se encuentra, no rompe el PDF

    c.setFont("Helvetica-Bold", 16)
    c.drawString(30, height - 40, "Alkimyk Food - Producción")
    c.setFont("Helvetica", 12)

    print("Tipo de PDF:", tipo)
    print("Canastos:", canastos)
    print("Detalles por sabor:", detalles_por_sabor)
    print("Total ingredientes:", total_ingredientes)

    if tipo == 'por_sabor':
        y = height - 50
        for sabor, ingredientes in detalles_por_sabor.items():
            cantidad_canastos = canastos[sabor]
            # Cálculo de cajas por sabor, especial para 'original'
            if sabor == 'original':
                cantidad_cajas = round((cantidad_canastos * 32) / 108)
            else:
                cantidad_cajas = round((cantidad_canastos * UNIDADES_POR_CANASTO) / (15 * 4))
            c.drawString(30, y, f"{sabor.replace('_', ' ').capitalize()} ({cantidad_canastos} canastos, {cantidad_cajas} cajas)")
            y -= 20
            for ingr, cant in ingredientes.items():
                unidad = "kg" if cant >= 1000 else "g"
                c.drawString(50, y, f"{ingr}: {round(cant / 1000, 2) if unidad == 'kg' else round(cant, 2)} {unidad}")
                y -= 15
                if y < 50:
                    c.showPage()
                    y = height - 50
            y -= 10  # Espacio entre sabores
    elif tipo == 'total':
        y = height - 50
        c.drawString(30, y, "Total de Ingredientes")
        y -= 20
        for ingr, cant in total_ingredientes.items():
            if ingr in ['Soja', 'Harina']:
                unidad = "kg"
                cantidad = round(cant, 2)
            else:
                if cant >= 1000:
                    unidad = "kg"
                    cantidad = round(cant / 1000, 2)
                else:
                    unidad = "g"
                    cantidad = round(cant, 2)
            c.drawString(50, y, f"{ingr}: {cantidad} {unidad}")
            y -= 15
            if y < 50:
                c.showPage()
                y = height - 50
        y -= 10  # Espacio antes del total
        c.drawString(30, y, f"Total de canastos: {total_canastos}")
        y -= 15
        if y < 50:
            c.showPage()
            y = height - 50
        c.drawString(30, y, f"Total de cajas: {total_cajas}")

    c.save()
    pdf_output.seek(0)
    return send_file(pdf_output, download_name='produccion_alkimyk.pdf', as_attachment=True)

def abrir_navegador():
    webbrowser.open("http://127.0.0.1:5000")


# Configuración de cupo diario y días de producción
@app.route('/configuracion', methods=['GET', 'POST'])
def configuracion():
    if request.method == 'POST':
        cupo_diario = request.form.get('cupo_diario')
        incluir_sabado = 'incluir_sabado' in request.form
        incluir_domingo = 'incluir_domingo' in request.form

        try:
            session['cupo_diario'] = int(cupo_diario)
        except:
            session['cupo_diario'] = 110  # valor por defecto si hay error

        session['incluir_sabado'] = incluir_sabado
        session['incluir_domingo'] = incluir_domingo

        # Capturar días de producción seleccionados por el usuario
        dias_habilitados = request.form.getlist('dias_habilitados')
        session['dias_habilitados'] = dias_habilitados if dias_habilitados else ['lunes', 'martes', 'miércoles', 'jueves', 'viernes']

        flash("Configuración guardada correctamente.", "success")
        return redirect(url_for('home'))

    # valores actuales o por defecto
    cupo_actual = session.get('cupo_diario', 110)
    sabado = session.get('incluir_sabado', False)
    domingo = session.get('incluir_domingo', False)
    return render_template(
        'configuracion.html',
        cupo_diario=cupo_actual,
        incluir_sabado=sabado,
        incluir_domingo=domingo,
        dias_habilitados=session.get('dias_habilitados', ['lunes', 'martes', 'miércoles', 'jueves', 'viernes'])
    )



# Ruta para login de administrador
@app.route('/login_admin', methods=['GET', 'POST'])
def login_admin():
    from werkzeug.security import check_password_hash
    if request.method == 'POST':
        email = request.form.get('email')
        password = request.form.get('password')
        user = Usuario.query.filter_by(email=email).first()
        if user and check_password_hash(user.password, password):
            session['usuario'] = user.nombre
            session['rol'] = user.rol
            if user.rol == 'admin':
                session['post_login_redirect'] = '/'
                flash('Inicio de sesión como administrador', 'success')
                return redirect('/splash')
            else:
                flash('Inicio de sesión exitoso', 'success')
                return redirect(url_for('home'))
        else:
            flash('Credenciales inválidas o sin permiso', 'danger')
    return render_template('login_admin.html')



# Ruta para crear un nuevo usuario (solo admin)
@app.route('/crear_usuario', methods=['GET', 'POST'])
def crear_usuario():
    if session.get('rol') != 'admin':
        flash('Acceso restringido solo para administradores.', 'danger')
        return redirect(url_for('home'))

    if request.method == 'POST':
        nombre = request.form.get('nombre')
        email = request.form.get('email')
        password = request.form.get('password')
        rol = request.form.get('rol')

        from werkzeug.security import generate_password_hash
        password_hash = generate_password_hash(password)

        if Usuario.query.filter_by(email=email).first():
            flash('El email ya está registrado.', 'danger')
        else:
            nuevo_usuario = Usuario(nombre=nombre, email=email, password=password_hash, rol=rol)
            db.session.add(nuevo_usuario)
            db.session.commit()
            flash('Usuario creado correctamente.', 'success')

    return render_template('crear_usuario.html')



# Ruta para cerrar sesión
@app.route('/logout')
def logout():
    session.clear()
    flash('Sesión cerrada correctamente.', 'success')
    return redirect(url_for('login_admin'))


# Nueva ruta /costos
@app.route('/costos')
def costos():
    canastos = {}
    if 'usuario' in session:
        usuario_email = session['usuario']
        producciones = Produccion.query.filter_by(usuario_email=usuario_email).all()
        for p in producciones:
            canastos[p.sabor] = canastos.get(p.sabor, 0) + p.canastos
    if not canastos:
        flash("No hay datos de producción cargados.", "warning")
        return redirect(url_for('canastos'))

    total_canastos = sum(canastos.values())
    total_ingredientes = {}

    def add(dic, nombre, cantidad_g):
        dic[nombre] = dic.get(nombre, 0) + cantidad_g

    # Solo para sabores que NO sean 'original'
    non_original_canastos = sum(c for s, c in canastos.items() if s != 'original')
    if non_original_canastos > 0:
        add(total_ingredientes, 'Soja', MASA_BASE_POR_100_CANASTOS['soja_kg'] * non_original_canastos / MASA_BASE_CANASTOS)
        add(total_ingredientes, 'Chimichurri', MASA_BASE_POR_100_CANASTOS['chimichurri_g'] * non_original_canastos / MASA_BASE_CANASTOS)
        add(total_ingredientes, 'Sal', MASA_BASE_POR_100_CANASTOS['sal_g'] * non_original_canastos / MASA_BASE_CANASTOS)

    for sabor, cantidad in canastos.items():
        if cantidad == 0:
            continue
        if sabor == 'original':
            receta_originales = {
                'Soja': 75,
                'Harina': 30,
                'Sal': 0.5,
                'Chimichurri': 0.8
            }
            for ingrediente, total_base in receta_originales.items():
                cantidad_ingrediente = (total_base / 85) * cantidad
                if ingrediente in ['Soja', 'Harina']:
                    cantidad_final = round(cantidad_ingrediente, 2)
                else:
                    cantidad_final = round(cantidad_ingrediente * 1000, 2) if cantidad_ingrediente < 1 else round(cantidad_ingrediente, 2)
                add(total_ingredientes, ingrediente, cantidad_final)
            # continuar, pero pan rallado se suma abajo
            continue
        unidades = cantidad * UNIDADES_POR_CANASTO
        add(total_ingredientes, 'Soja', MASA_BASE_POR_100_CANASTOS['soja_kg'] * cantidad / MASA_BASE_CANASTOS)
        add(total_ingredientes, 'Chimichurri', MASA_BASE_POR_100_CANASTOS['chimichurri_g'] * cantidad / MASA_BASE_CANASTOS)
        add(total_ingredientes, 'Sal', MASA_BASE_POR_100_CANASTOS['sal_g'] * cantidad / MASA_BASE_CANASTOS)
        if sabor == 'aceituna':
            add(total_ingredientes, 'Muzzarella', unidades * 15)
            add(total_ingredientes, 'Aceitunas', unidades * 20)
        elif sabor == 'caprese':
            tomate_total = unidades * 25
            add(total_ingredientes, 'Muzzarella', unidades * 15)
            add(total_ingredientes, 'Tomate', tomate_total)
            add(total_ingredientes, 'Albahaca', unidades * 2)
            add(total_ingredientes, 'Sal', (tomate_total / 1000) * 4)
        elif sabor == 'queso_azul':
            mezcla_total = unidades * 30
            porc_queso = 2.3 / (18 + 2.3)
            porc_muzza = 1 - porc_queso
            add(total_ingredientes, 'Muzzarella', mezcla_total * porc_muzza)
            add(total_ingredientes, 'Queso Azul', mezcla_total * porc_queso)
        elif sabor == 'cebolla':
            cebolla_cruda = (unidades * 40) / 0.8
            add(total_ingredientes, 'Cebolla', cebolla_cruda)
            add(total_ingredientes, 'Orégano', (cebolla_cruda / 1000) * 2)
            add(total_ingredientes, 'Sal', (cebolla_cruda / 1000) * 5)
        elif sabor == 'espinaca':
            total_relleno = unidades * 40 / 0.9
            espinaca = total_relleno * 0.5 / 0.9
            cebolla = total_relleno * 0.25 / 0.8
            morron = total_relleno * 0.25 / 0.8
            add(total_ingredientes, 'Espinaca', espinaca)
            add(total_ingredientes, 'Cebolla', cebolla)
            add(total_ingredientes, 'Morrón', morron)
            add(total_ingredientes, 'Nuez Moscada', total_relleno / 1000 * 1)
            add(total_ingredientes, 'Pimienta Negra', total_relleno / 1000 * 1)
            add(total_ingredientes, 'Sal', total_relleno / 1000 * 5)
        elif sabor == 'calabaza':
            total_relleno = unidades * 40 / 0.8
            add(total_ingredientes, 'Calabaza', total_relleno)
            add(total_ingredientes, 'Cúrcuma', total_relleno / 1000 * 5)
            add(total_ingredientes, 'Sal', total_relleno / 1000 * 5)
        elif sabor == 'brocoli':
            total_relleno = unidades * 40
            add(total_ingredientes, 'Brócoli', total_relleno * 0.6)
            add(total_ingredientes, 'Cebolla', total_relleno * 0.4 / 0.8)
            add(total_ingredientes, 'Chimichurri', total_relleno / 1000 * 5)
            add(total_ingredientes, 'Sal', total_relleno / 1000 * 5)

    # Agregar pan rallado (10 g por unidad)
    for sabor, cantidad in canastos.items():
        if cantidad == 0:
            continue
        unidades_por_canasto = 32 if sabor == 'original' else UNIDADES_POR_CANASTO
        total_unidades = cantidad * unidades_por_canasto
        gramos_pan_rallado = total_unidades * 10
        total_ingredientes['Pan Rallado'] = total_ingredientes.get('Pan Rallado', 0) + gramos_pan_rallado

    # Obtener costos fijos desde la base de datos para el usuario
    costos_fijos = {}
    if 'usuario' in session:
        usuario_email = session['usuario']
        costos_fijos_query = CostoFijo.query.filter_by(usuario_email=usuario_email).all()
        costos_fijos = {c.nombre: c.monto for c in costos_fijos_query}

    # Obtener precios de ingredientes previos si existen (claves normalizadas)
    precios_ingredientes = {}
    if 'usuario' in session:
        usuario_email = session['usuario']
        precios = PrecioIngrediente.query.filter_by(usuario_email=usuario_email).all()
        for p in precios:
            clave_limpia = slugify(p.ingrediente)
            precios_ingredientes[clave_limpia] = p.precio_unitario

    # Obtener precios de venta por sabor desde la base si están disponibles
    precios_venta_por_sabor = {}
    if 'usuario' in session:
        usuario_email = session['usuario']
        precios_db = PrecioVentaSabor.query.filter_by(usuario_email=usuario_email).all()
        for p in precios_db:
            precios_venta_por_sabor[p.sabor] = p.precio

    return render_template(
        'costos.html',
        ingredientes=total_ingredientes,
        costos_fijos=costos_fijos,
        precios_costos_fijos=costos_fijos,
        precios_ingredientes=precios_ingredientes,
        precios_venta_por_sabor=precios_venta_por_sabor,
        precios_venta=precios_venta_por_sabor
    )


# Ruta para guardar precios unitarios de ingredientes
@app.route('/guardar_precios_ingredientes', methods=['POST'])
def guardar_precios_ingredientes():
    if 'usuario' not in session:
        return jsonify({'success': False, 'message': 'Usuario no autenticado'})

    usuario_email = session['usuario']
    data = request.get_json()

    if not data:
        return jsonify({'success': False, 'message': 'No se recibieron datos'})

    # Borrar los precios anteriores del usuario
    PrecioIngrediente.query.filter_by(usuario_email=usuario_email).delete()

    for ingrediente, precio in data.items():
        ingrediente_limpio = slugify(ingrediente)
        precio_unitario = normalizar_importe(precio)
        if precio_unitario > 0:
            nuevo_precio = PrecioIngrediente(usuario_email=usuario_email, ingrediente=ingrediente_limpio, precio_unitario=precio_unitario)
            db.session.add(nuevo_precio)

    db.session.commit()
    return jsonify({'success': True, 'message': 'Precios guardados correctamente'})

@app.route('/dashboard_rentabilidad')
def dashboard_rentabilidad():
    from flask import jsonify
    # Definir slugify si no está visible
    def slugify(nombre):
        return nombre.strip().lower().replace(' ', '_').replace('á', 'a').replace('é', 'e').replace('í', 'i').replace('ó', 'o').replace('ú', 'u').replace('ñ', 'n')

    canastos = session.get('canastos', {})
    detalles_por_sabor = {}

    # Traer los precios unitarios desde la base de datos
    precios_ingredientes = {}
    if 'usuario' in session:
        usuario_email = session['usuario']
        precios = PrecioIngrediente.query.filter_by(usuario_email=usuario_email).all()
        for p in precios:
            precios_ingredientes[p.ingrediente] = p.precio_unitario

    for sabor, cantidad in canastos.items():
        if cantidad == 0:
            continue
        temp = {}
        if sabor == 'original':
            receta_originales = {
                'Soja': 75,
                'Harina': 30,
                'Sal': 0.5,
                'Chimichurri': 0.8
            }
            for ingrediente, total_base in receta_originales.items():
                cantidad_ingrediente = (total_base / 85) * cantidad
                if ingrediente in ['Soja', 'Harina']:
                    cantidad_final = round(cantidad_ingrediente, 2)
                else:
                    cantidad_final = round(cantidad_ingrediente * 1000, 2) if cantidad_ingrediente < 1 else round(cantidad_ingrediente, 2)
                temp[ingrediente] = cantidad_final
        else:
            unidades = cantidad * UNIDADES_POR_CANASTO
            temp['Soja'] = MASA_BASE_POR_100_CANASTOS['soja_kg'] * cantidad / MASA_BASE_CANASTOS
            temp['Chimichurri'] = MASA_BASE_POR_100_CANASTOS['chimichurri_g'] * cantidad / MASA_BASE_CANASTOS
            temp['Sal'] = MASA_BASE_POR_100_CANASTOS['sal_g'] * cantidad / MASA_BASE_CANASTOS
            if sabor == 'aceituna':
                temp['Muzzarella'] = unidades * 15
                temp['Aceitunas'] = unidades * 20
            elif sabor == 'caprese':
                tomate_total = unidades * 25
                temp['Muzzarella'] = unidades * 15
                temp['Tomate'] = tomate_total
                temp['Albahaca'] = unidades * 2
                temp['Sal'] = temp.get('Sal', 0) + (tomate_total / 1000) * 4
            elif sabor == 'queso_azul':
                mezcla_total = unidades * 30
                porc_queso = 2.3 / (18 + 2.3)
                porc_muzza = 1 - porc_queso
                temp['Muzzarella'] = mezcla_total * porc_muzza
                temp['Queso Azul'] = mezcla_total * porc_queso
            elif sabor == 'cebolla':
                cebolla_cruda = (unidades * 40) / 0.8
                temp['Cebolla'] = cebolla_cruda
                temp['Orégano'] = (cebolla_cruda / 1000) * 2
                temp['Sal'] = temp.get('Sal', 0) + (cebolla_cruda / 1000) * 5
            elif sabor == 'espinaca':
                total_relleno = unidades * 40 / 0.9
                espinaca = total_relleno * 0.5 / 0.9
                cebolla = total_relleno * 0.25 / 0.8
                morron = total_relleno * 0.25 / 0.8
                temp['Espinaca'] = espinaca
                temp['Cebolla'] = cebolla
                temp['Morrón'] = morron
                temp['Nuez Moscada'] = total_relleno / 1000 * 1
                temp['Pimienta Negra'] = total_relleno / 1000 * 1
                temp['Sal'] = temp.get('Sal', 0) + total_relleno / 1000 * 5
            elif sabor == 'calabaza':
                total_relleno = unidades * 40 / 0.8
                temp['Calabaza'] = total_relleno
                temp['Cúrcuma'] = total_relleno / 1000 * 5
                temp['Sal'] = temp.get('Sal', 0) + total_relleno / 1000 * 5
            elif sabor == 'brocoli':
                total_relleno = unidades * 40
                temp['Brócoli'] = total_relleno * 0.6
                temp['Cebolla'] = total_relleno * 0.4 / 0.8
                temp['Chimichurri'] = total_relleno / 1000 * 5
                temp['Sal'] = temp.get('Sal', 0) + total_relleno / 1000 * 5
        detalles_por_sabor[sabor] = temp
    # Calcular el costo total por sabor usando slugify en la clave
    for sabor in detalles_por_sabor:
        costo_total_sabor = 0
        for ingr, cant in detalles_por_sabor[sabor].items():
            clave_limpia = slugify(ingr)
            precio_unitario = precios_ingredientes.get(clave_limpia, 0)
            # Para cantidades >= 1000, se asume kg, si no gramos, pero el precio es por kg
            costo_total_sabor += (cant / 1000) * precio_unitario if cant >= 0 else 0
        detalles_por_sabor[sabor]['Costo Variable Total'] = costo_total_sabor
        print(f"Receta para {sabor}: {detalles_por_sabor[sabor]}")

    # Agregar pan rallado (10 g por unidad)
    for sabor, cantidad in canastos.items():
        if cantidad == 0:
            continue
        unidades_por_canasto = 32 if sabor == 'original' else UNIDADES_POR_CANASTO
        total_unidades = cantidad * unidades_por_canasto
        gramos_pan_rallado = total_unidades * 10
        detalles_por_sabor.setdefault(sabor, {})
        detalles_por_sabor[sabor]['Pan Rallado'] = gramos_pan_rallado

    # Calcular packaging por sabor
    total_packaging_por_sabor = {}
    try:
        usuario_email = session['usuario']
        precios_ingredientes_db = PrecioIngrediente.query.filter_by(usuario_email=usuario_email).all()
        precios_map = {p.ingrediente.lower(): p.precio_unitario for p in precios_ingredientes_db}
        precio_pack = precios_map.get('packaging', 0)
        precio_caja = precios_map.get('cajas', 0)
    except:
        precio_pack = 0
        precio_caja = 0
    for sabor, cantidad_canastos in canastos.items():
        if cantidad_canastos == 0:
            continue
        # Cálculo correcto de total_unidades según sabor
        if sabor == 'original':
            total_unidades = cantidad_canastos * 32
            total_packs = total_unidades / 4
            cantidad_cajas = total_unidades / 108
        else:
            total_unidades = cantidad_canastos * UNIDADES_POR_CANASTO
            total_packs = total_unidades / 4
            cantidad_cajas = round(total_packs / 15)
        costo_packaging = total_packs * precio_pack + cantidad_cajas * precio_caja
        total_packaging_por_sabor[sabor] = costo_packaging

    # Obtener costos fijos para el usuario
    costos_fijos = {}
    if 'usuario' in session:
        usuario_email = session['usuario']
        costos_fijos_query = CostoFijo.query.filter_by(usuario_email=usuario_email).all()
        costos_fijos = {c.nombre: c.monto for c in costos_fijos_query}

    # Agregar variable index=True al contexto antes del render_template
    return render_template(
        'dashboard_rentabilidad.html',
        detalles_por_sabor=detalles_por_sabor,
        total_packaging_por_sabor=total_packaging_por_sabor,
        costos_fijos=costos_fijos,
        precios_ingredientes=precios_ingredientes,
        canastos=canastos,
        index=True
    )

@app.route('/resumen_datos')
def resumen_datos():
    canastos = session.get('canastos', {})
    total_canastos = sum(canastos.values())
    total_cajas = round((total_canastos * UNIDADES_POR_CANASTO) / (15 * 4))
    total_ingredientes = {}

    def add(dic, nombre, cantidad_g):
        dic[nombre] = dic.get(nombre, 0) + cantidad_g

    add(total_ingredientes, 'Soja', MASA_POR_94_CANASTOS['soja_kg'] * total_canastos / CANASTOS_BASE)
    add(total_ingredientes, 'Harina', MASA_POR_94_CANASTOS['harina_kg'] * total_canastos / CANASTOS_BASE)
    add(total_ingredientes, 'Chimichurri', MASA_POR_94_CANASTOS['chimichurri_g'] * total_canastos / CANASTOS_BASE)
    add(total_ingredientes, 'Sal', MASA_POR_94_CANASTOS['sal_g'] * total_canastos / CANASTOS_BASE)

    detalles_por_sabor = {}
    for sabor, cantidad in canastos.items():
        if cantidad == 0:
            continue
        unidades = cantidad * UNIDADES_POR_CANASTO
        temp = {}
        if sabor == 'aceituna':
            temp['Muzzarella'] = unidades * 15
            temp['Aceitunas'] = unidades * 20
        elif sabor == 'caprese':
            tomate_total = unidades * 25
            temp['Muzzarella'] = unidades * 15
            temp['Tomate'] = tomate_total
            temp['Albahaca'] = unidades * 2
            temp['Sal'] = (tomate_total / 1000) * 4
        elif sabor == 'queso_azul':
            mezcla_total = unidades * 30
            porc_queso = 2.3 / (18 + 2.3)
            porc_muzza = 1 - porc_queso
            temp['Muzzarella'] = mezcla_total * porc_muzza
            temp['Queso Azul'] = mezcla_total * porc_queso
        elif sabor == 'cebolla':
            cebolla_cruda = (unidades * 40) / 0.8
            temp['Cebolla'] = cebolla_cruda
            temp['Orégano'] = (cebolla_cruda / 1000) * 2
            temp['Sal'] = (cebolla_cruda / 1000) * 5
        elif sabor == 'espinaca':
            total_relleno = unidades * 40 / 0.9
            espinaca = total_relleno * 0.5 / 0.9
            cebolla = total_relleno * 0.25 / 0.8
            morron = total_relleno * 0.25 / 0.8
            temp['Espinaca'] = espinaca
            temp['Cebolla'] = cebolla
            temp['Morrón'] = morron
            temp['Nuez Moscada'] = total_relleno / 1000 * 1
            temp['Pimienta Negra'] = total_relleno / 1000 * 1
            temp['Sal'] = total_relleno / 1000 * 5
        elif sabor == 'calabaza':
            total_relleno = unidades * 40 / 0.8
            temp['Calabaza'] = total_relleno
            temp['Cúrcuma'] = total_relleno / 1000 * 5
            temp['Sal'] = total_relleno / 1000 * 5
        elif sabor == 'brocoli':
            total_relleno = unidades * 40
            temp['Brócoli'] = total_relleno * 0.6
            temp['Cebolla'] = total_relleno * 0.4 / 0.8
            temp['Chimichurri'] = total_relleno / 1000 * 5
            temp['Sal'] = total_relleno / 1000 * 5
        detalles_por_sabor[sabor] = temp
        for k, v in temp.items():
            add(total_ingredientes, k, v)

    # Agregar pan rallado (10 g por unidad)
    for sabor, cantidad in canastos.items():
        if cantidad == 0:
            continue
        unidades_por_canasto = 32 if sabor == 'original' else UNIDADES_POR_CANASTO
        total_unidades = cantidad * unidades_por_canasto
        gramos_pan_rallado = total_unidades * 10
        detalles_por_sabor.setdefault(sabor, {})
        detalles_por_sabor[sabor]['Pan Rallado'] = gramos_pan_rallado
        add(total_ingredientes, 'Pan Rallado', gramos_pan_rallado)

    # Agregar masa base (soja, harina, chimichurri, sal) a detalles_por_sabor para cada sabor
    receta_masa_por_canasto = {
        'Soja': MASA_POR_94_CANASTOS['soja_kg'] / CANASTOS_BASE,
        'Harina': MASA_POR_94_CANASTOS['harina_kg'] / CANASTOS_BASE,
        'Chimichurri': MASA_POR_94_CANASTOS['chimichurri_g'] / CANASTOS_BASE,
        'Sal': MASA_POR_94_CANASTOS['sal_g'] / CANASTOS_BASE
    }
    for sabor, cantidad_canastos in canastos.items():
        if sabor not in detalles_por_sabor:
            detalles_por_sabor[sabor] = {}
        for ingrediente, cantidad_por_canasto in receta_masa_por_canasto.items():
            detalles_por_sabor[sabor][ingrediente] = detalles_por_sabor[sabor].get(ingrediente, 0) + (cantidad_por_canasto * cantidad_canastos)

    # Agregar datos faltantes desde la base de datos
    usuario_email = session.get('usuario')
    ingredientes_db = PrecioIngrediente.query.filter_by(usuario_email=usuario_email).all()
    precios_dict = {i.ingrediente: i.precio_unitario for i in ingredientes_db}
    
    costos_fijos_db = CostoFijo.query.filter_by(usuario_email=usuario_email).all()
    costos_fijos_dict = {c.nombre: c.monto for c in costos_fijos_db}
    
    precios_venta_db = PrecioVentaSabor.query.filter_by(usuario_email=usuario_email).all()
    precios_venta_dict = {p.sabor: p.precio for p in precios_venta_db}

    return jsonify({
        'canastos': canastos,
        'total_canastos': total_canastos,
        'total_cajas': total_cajas,
        'ingredientes_totales': total_ingredientes,
        'detalles_por_sabor': detalles_por_sabor,
        'precios_ingredientes': precios_dict,
        'costos_fijos': costos_fijos_dict,
        'precios_venta': precios_venta_dict,
        'packaging': precios_dict.get("packaging", 0),
        'cajas': precios_dict.get("cajas", 0)
    })


# Ruta para /splash
@app.route('/splash')
def splash():
    return render_template('splash.html')



# Crear tablas si no existen antes de cada petición (garantiza inicialización en todos los entornos)
@app.before_request
def crear_tablas_si_no_existen():
    db.create_all()
    # Crear usuario administrador por defecto si no existe
    from werkzeug.security import generate_password_hash
    if not Usuario.query.filter_by(email='alkimykfood@gmail.com').first():
        admin = Usuario(
            nombre='Administrador',
            email='alkimykfood@gmail.com',
            password=generate_password_hash('Mica1979'),
            rol='admin'
        )
        db.session.add(admin)
        db.session.commit()


# Ruta para guardar costos fijos del usuario
@app.route('/guardar_costos', methods=['POST'])
def guardar_costos():
    if 'usuario' not in session:
        return jsonify({'success': False, 'message': 'Usuario no autenticado'})

    usuario_email = session['usuario']
    data = request.get_json()

    if not data:
        return jsonify({'success': False, 'message': 'No se recibieron datos'})

    # Borrar los costos fijos anteriores del usuario
    CostoFijo.query.filter_by(usuario_email=usuario_email).delete()

    for nombre, monto in data.items():
        monto_float = normalizar_importe(monto)
        nuevo_costo = CostoFijo(usuario_email=usuario_email, nombre=nombre, monto=monto_float)
        db.session.add(nuevo_costo)

    db.session.commit()
    return jsonify({'success': True, 'message': 'Costos guardados correctamente'})



# Ruta para guardar resumen histórico de rentabilidad
@app.route('/guardar_resumen_historico', methods=['POST'])
def guardar_resumen_historico():
    if 'usuario' not in session:
        return jsonify({'success': False, 'message': 'Usuario no autenticado'})

    data = request.get_json()
    print("📥 Datos recibidos para histórico:", data)
    usuario_email = session['usuario']
    try:
        # Controlar campos opcionales
        total_canastos = data.get('total_canastos') or 0
        total_cajas = data.get('total_cajas') or 0
        total_facturar = data.get('total_facturar') or 0
        total_con_iva = data.get('total_con_iva') or 0
        ganancia_total = data.get('ganancia_total') or 0
        rentabilidad = data.get('rentabilidad') or 0
        nuevo = ResumenHistorico(
            usuario_email=usuario_email,
            total_canastos=total_canastos,
            total_cajas=total_cajas,
            total_facturar=total_facturar,
            total_con_iva=total_con_iva,
            ganancia_total=ganancia_total,
            rentabilidad=rentabilidad
        )
        db.session.add(nuevo)
        db.session.commit()
        print("✅ Histórico guardado:", nuevo)
        # Devolver fecha para confirmación
        return jsonify({'success': True, 'fecha': nuevo.fecha.strftime('%Y-%m-%d')})
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)})

def slugify(nombre):
    return nombre.strip().lower().replace(' ', '_').replace('á', 'a').replace('é', 'e').replace('í', 'i').replace('ó', 'o').replace('ú', 'u').replace('ñ', 'n')

@app.route('/guardar_todos_los_costos', methods=['POST'])
def guardar_todos_los_costos():
    if 'usuario' not in session:
        return jsonify({'success': False, 'message': 'Usuario no autenticado'})

    # Importar el formateador de números para consola
    from babel.numbers import format_decimal

    usuario_email = session['usuario']
    data = request.get_json()
    print("🔍 Datos recibidos en /guardar_todos_los_costos:", data)
    costos_fijos = data.get('costos_fijos', {})
    precios_venta = data.get('precios_venta', {})

    print("💰 Costos fijos:")
    for nombre, val in costos_fijos.items():
        try:
            val_float = float(val)
            val_format = format_decimal(val_float, locale='es_AR')
        except:
            val_format = val
        print(f"  - {nombre}: {val_format}")

    print("🏷️ Precios de venta:")
    for sabor, val in precios_venta.items():
        try:
            val_float = float(val)
            val_format = format_decimal(val_float, locale='es_AR')
        except:
            val_format = val
        print(f"  - {sabor}: {val_format}")

    precios_ingredientes = data.get('ingredientes', {})
    print("📦 Ingredientes recibidos:")
    for ingr, val in precios_ingredientes.items():
        try:
            val_float = float(val)
            val_format = format_decimal(val_float, locale='es_AR')
        except:
            val_format = val
        print(f"  - {ingr}: {val_format}")

    if not data:
        return jsonify({'success': False, 'message': 'No se recibieron datos'})

    # Validar que ningún precio de ingrediente esté vacío o igual a 0 antes de guardar
    for ingrediente, precio in precios_ingredientes.items():
        if precio is None or str(precio).strip() == "" or normalizar_importe(precio) == 0:
            return jsonify({'success': False, 'message': f'Debes completar un valor mayor a 0 para el ingrediente: {ingrediente}'})

    # Guardar precios de ingredientes con normalización robusta
    PrecioIngrediente.query.filter_by(usuario_email=usuario_email).delete()
    for ingrediente, precio in precios_ingredientes.items():
        ingrediente_limpio = slugify(ingrediente)
        precio_unitario = normalizar_importe(precio)
        if precio_unitario <= 0:
            return jsonify({'success': False, 'message': f'Debes completar un valor mayor a 0 para el ingrediente: {ingrediente}'})
        nuevo_precio = PrecioIngrediente(usuario_email=usuario_email, ingrediente=ingrediente_limpio, precio_unitario=precio_unitario)
        db.session.add(nuevo_precio)

    # Guardar costos fijos
    CostoFijo.query.filter_by(usuario_email=usuario_email).delete()
    for nombre, monto in costos_fijos.items():
        monto_float = normalizar_importe(monto)
        nuevo_costo = CostoFijo(usuario_email=usuario_email, nombre=nombre, monto=monto_float)
        db.session.add(nuevo_costo)

    # Guardar precios de venta en la base de datos (normalización robusta)
    PrecioVentaSabor.query.filter_by(usuario_email=usuario_email).delete()
    for sabor, precio_str in precios_venta.items():
        precio_float = normalizar_importe(precio_str)
        nuevo_precio = PrecioVentaSabor(usuario_email=usuario_email, sabor=sabor, precio=precio_float)
        db.session.add(nuevo_precio)

    db.session.commit()
    return jsonify({'success': True, 'message': 'Todos los datos guardados correctamente'})




@app.route('/planificacion')
def planificacion():
    if 'usuario' not in session:
        flash("Debe iniciar sesión para acceder a la planificación", "warning")
        return redirect(url_for('login_admin'))

    canastos = session.get('canastos', {})
    total_canastos = sum(canastos.values())

    # Calcular total de cajas y cajas por sabor
    total_cajas = 0
    cajas_por_sabor = {}
    for sabor, cantidad in canastos.items():
        unidades_por_canasto = 32 if sabor == 'original' else UNIDADES_POR_CANASTO
        total_unidades = cantidad * unidades_por_canasto
        cajas = round(total_unidades / (108 if sabor == 'original' else (15 * 4)))
        total_cajas += cajas
        cajas_por_sabor[sabor] = cajas

    # Verificar si se cargó previamente total_ingredientes_fmt desde /canastos
    if 'total_ingredientes_fmt' not in session:
        flash("Faltan datos de ingredientes. Por favor completá los canastos primero.", "warning")
        return redirect(url_for('canastos'))

    total_ingredientes_fmt = session['total_ingredientes_fmt']

    return render_template('planificacion.html',
                           canastos=canastos,
                           total_canastos=total_canastos,
                           total_cajas=total_cajas,
                           cajas_por_sabor=cajas_por_sabor,
                           total_ingredientes_fmt=total_ingredientes_fmt)
from flask import render_template, request, redirect, url_for, flash
from app import app, db  # asegurate que app y db estén importados correctamente
from datetime import datetime


# Lista fija de sabores disponibles
SABORES_DISPONIBLES = [
    'Caprese',
    'Aceituna',
    'Queso Azul',
    'Cebolla',
    'Espinaca',
    'Calabaza',
    'Brócoli',
    'Original'
]

@app.route('/produccion_diaria', methods=['GET', 'POST'])
def produccion_diaria():
    if request.method == 'POST':
        try:
            fecha = datetime.strptime(request.form['fecha'], '%Y-%m-%d').date()
            sabor = request.form['sabor']
            cantidad = int(request.form['cantidad'])

            nueva = ProduccionDiaria(fecha=fecha, sabor=sabor, cantidad_canastos=cantidad)
            db.session.add(nueva)
            db.session.commit()
            flash("✅ Producción guardada correctamente.", "success")
        except Exception as e:
            db.session.rollback()
            flash(f"❌ Error al guardar: {str(e)}", "danger")

        return redirect(url_for('produccion_diaria'))

    # Filtro por fechas
    fecha_inicio = request.args.get('fecha_inicio')
    fecha_fin = request.args.get('fecha_fin')
    producciones = []

    if fecha_inicio and fecha_fin:
        try:
            f1 = datetime.strptime(fecha_inicio, '%Y-%m-%d').date()
            f2 = datetime.strptime(fecha_fin, '%Y-%m-%d').date()
            producciones = ProduccionDiaria.query.filter(
                ProduccionDiaria.fecha.between(f1, f2)
            ).order_by(ProduccionDiaria.fecha).all()
        except Exception:
            flash("❗Formato de fecha inválido", "warning")

    return render_template(
        'produccion_diaria.html',
        producciones=producciones,
        sabores=SABORES_DISPONIBLES
    )
@app.route('/produccion_resultado')
def produccion_resultado():
    from datetime import datetime
    fecha_inicio = request.args.get('fecha_inicio')
    fecha_fin = request.args.get('fecha_fin')

    if not fecha_inicio or not fecha_fin:
        flash("Faltan fechas para la búsqueda", "warning")
        return redirect(url_for('produccion_diaria'))

    f1 = datetime.strptime(fecha_inicio, '%Y-%m-%d').date()
    f2 = datetime.strptime(fecha_fin, '%Y-%m-%d').date()

    registros = ProduccionDiaria.query.filter(
        db.func.date(ProduccionDiaria.fecha).between(f1, f2)
    ).order_by(ProduccionDiaria.fecha).all()

    dias_agrupados = {}

    dias_es = {
        'Monday': 'Lunes',
        'Tuesday': 'Martes',
        'Wednesday': 'Miércoles',
        'Thursday': 'Jueves',
        'Friday': 'Viernes',
        'Saturday': 'Sábado',
        'Sunday': 'Domingo'
    }

    for r in registros:
        fecha_str = r.fecha.strftime('%Y-%m-%d')
        dia_semana_en = r.fecha.strftime('%A')
        dia_semana = dias_es.get(dia_semana_en, dia_semana_en)

        if fecha_str not in dias_agrupados:
            dias_agrupados[fecha_str] = {
                'fecha': fecha_str,
                'dia_semana': dia_semana,
                'sabores': {}
            }

        if r.sabor not in dias_agrupados[fecha_str]['sabores']:
            dias_agrupados[fecha_str]['sabores'][r.sabor] = {
                'canastos': 0,
                'unidades': 0,
                'cajas': 0,
                'packs_sueltos': 0
            }

        dias_agrupados[fecha_str]['sabores'][r.sabor]['canastos'] += r.cantidad_canastos
        total_unidades = dias_agrupados[fecha_str]['sabores'][r.sabor]['canastos'] * 18
        cajas = total_unidades // 60
        sobrantes = total_unidades % 60
        packs_sueltos = sobrantes // 4

        dias_agrupados[fecha_str]['sabores'][r.sabor]['unidades'] = total_unidades
        dias_agrupados[fecha_str]['sabores'][r.sabor]['cajas'] = cajas
        dias_agrupados[fecha_str]['sabores'][r.sabor]['packs_sueltos'] = packs_sueltos

    # ✅ Agregado para mostrar total de canastos del día
    for dia in dias_agrupados.values():
        dia['total_canastos'] = sum(sabor['canastos'] for sabor in dia['sabores'].values())

    calendario = list(dias_agrupados.values())

    return render_template(
        'produccion_resultado.html',
        calendario=calendario,
        fecha_inicio=fecha_inicio,
        fecha_fin=fecha_fin
    )
@app.route('/eliminar_produccion_dia', methods=['POST'])
def eliminar_produccion_dia():
    from datetime import datetime
    fecha_str = request.form.get('fecha')
    fecha_inicio = request.form.get('fecha_inicio')
    fecha_fin = request.form.get('fecha_fin')

    try:
        fecha_obj = datetime.strptime(fecha_str, '%Y-%m-%d').date()
        ProduccionDiaria.query.filter(
            db.func.date(ProduccionDiaria.fecha) == fecha_obj
        ).delete()
        db.session.commit()
        flash(f'Producción del {fecha_str} eliminada correctamente.', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'Error al eliminar la producción: {str(e)}', 'danger')

    return redirect(url_for('produccion_resultado', fecha_inicio=fecha_inicio, fecha_fin=fecha_fin))